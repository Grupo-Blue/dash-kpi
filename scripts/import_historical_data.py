#!/usr/bin/env python3
"""
Script para importar dados históricos de KPIs a partir de planilha Excel
Popula a tabela kpiSnapshots no banco de dados
"""

import sys
import os
from openpyxl import load_workbook
from datetime import datetime
import mysql.connector
from mysql.connector import Error
import json

# Adicionar o diretório raiz ao path para importar módulos
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

def get_db_connection():
    """Conecta ao banco de dados MySQL usando DATABASE_URL"""
    database_url = os.getenv('DATABASE_URL')
    
    if not database_url:
        raise Exception("DATABASE_URL não encontrada nas variáveis de ambiente")
    
    # Parse DATABASE_URL: mysql://user:password@host:port/database
    # Formato esperado: mysql://username:password@host:port/database
    if database_url.startswith('mysql://'):
        database_url = database_url.replace('mysql://', '')
    
    parts = database_url.split('@')
    if len(parts) != 2:
        raise Exception(f"DATABASE_URL inválida: {database_url}")
    
    user_pass = parts[0].split(':')
    host_db = parts[1].split('/')
    
    if len(user_pass) != 2 or len(host_db) != 2:
        raise Exception(f"DATABASE_URL mal formatada: {database_url}")
    
    username = user_pass[0]
    password = user_pass[1]
    
    host_port = host_db[0].split(':')
    host = host_port[0]
    port = int(host_port[1]) if len(host_port) > 1 else 3306
    database = host_db[1].split('?')[0]  # Remove query params se houver
    
    try:
        connection = mysql.connector.connect(
            host=host,
            port=port,
            user=username,
            password=password,
            database=database
        )
        print(f"✅ Conectado ao banco de dados: {database}")
        return connection
    except Error as e:
        print(f"❌ Erro ao conectar ao banco de dados: {e}")
        raise

def insert_snapshot(cursor, company_id, snapshot_date, kpi_type, source, data):
    """Insere um snapshot no banco de dados"""
    query = """
    INSERT INTO kpiSnapshots (companyId, snapshotDate, kpiType, source, data, createdAt)
    VALUES (%s, %s, %s, %s, %s, NOW())
    """
    
    data_json = json.dumps(data)
    
    try:
        cursor.execute(query, (company_id, snapshot_date, kpi_type, source, data_json))
        return True
    except Error as e:
        print(f"❌ Erro ao inserir snapshot: {e}")
        return False

def import_blue_consult(wb, cursor):
    """Importa dados da aba Blue Consult"""
    print("\n📊 Importando Blue Consult...")
    
    if "Blue Consult" not in wb.sheetnames:
        print("⚠️  Aba 'Blue Consult' não encontrada")
        return 0
    
    ws = wb["Blue Consult"]
    count = 0
    
    # Pular cabeçalho (linha 1) e linha de exemplo (linha 2)
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Verificar se a linha está vazia
        if not row[0] or row[0] == "":
            continue
        
        try:
            snapshot_date = datetime.strptime(str(row[0]), "%Y-%m-%d")
            
            data = {
                "faturamento_mensal": float(row[1]) if row[1] else 0,
                "novos_clientes": int(row[2]) if row[2] else 0,
                "clientes_implantacao": int(row[3]) if row[3] else 0,
                "taxa_conversao": float(row[4]) if row[4] else 0,
                "receitas_nibo": float(row[5]) if row[5] else 0,
                "despesas_nibo": float(row[6]) if row[6] else 0,
                "saldo_nibo": float(row[7]) if row[7] else 0,
            }
            
            if insert_snapshot(cursor, 1, snapshot_date, 'blue_consult_all', 'consolidated', data):
                count += 1
                print(f"  ✓ {snapshot_date.strftime('%Y-%m-%d')}: Faturamento R$ {data['faturamento_mensal']:,.2f}")
        
        except Exception as e:
            print(f"  ✗ Erro na linha: {row[0]} - {e}")
    
    return count

def import_tokeniza_academy(wb, cursor):
    """Importa dados da aba Tokeniza Academy"""
    print("\n📊 Importando Tokeniza Academy...")
    
    if "Tokeniza Academy" not in wb.sheetnames:
        print("⚠️  Aba 'Tokeniza Academy' não encontrada")
        return 0
    
    ws = wb["Tokeniza Academy"]
    count = 0
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0] or row[0] == "":
            continue
        
        try:
            snapshot_date = datetime.strptime(str(row[0]), "%Y-%m-%d")
            
            data = {
                "total_membros_discord": int(row[1]) if row[1] else 0,
                "membros_online": int(row[2]) if row[2] else 0,
                "novos_membros_7d": int(row[3]) if row[3] else 0,
                "novos_membros_30d": int(row[4]) if row[4] else 0,
                "total_alunos_cademi": int(row[5]) if row[5] else 0,
                "alunos_ativos": int(row[6]) if row[6] else 0,
                "total_cursos": int(row[7]) if row[7] else 0,
            }
            
            if insert_snapshot(cursor, 4, snapshot_date, 'tokeniza_academy_all', 'consolidated', data):
                count += 1
                print(f"  ✓ {snapshot_date.strftime('%Y-%m-%d')}: {data['total_membros_discord']} membros Discord")
        
        except Exception as e:
            print(f"  ✗ Erro na linha: {row[0]} - {e}")
    
    return count

def import_metricool(wb, cursor):
    """Importa dados da aba Redes Sociais"""
    print("\n📊 Importando Redes Sociais (Metricool)...")
    
    if "Redes Sociais" not in wb.sheetnames:
        print("⚠️  Aba 'Redes Sociais' não encontrada")
        return 0
    
    ws = wb["Redes Sociais"]
    count = 0
    
    # Mapeamento de empresas para IDs
    company_map = {
        "Blue Consult": 1,
        "Tokeniza": 2,
        "Tokeniza Academy": 4,
        "Mychel Mendes": 30004,
    }
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0] or row[0] == "" or not row[1]:
            continue
        
        try:
            snapshot_date = datetime.strptime(str(row[0]), "%Y-%m-%d")
            company_name = str(row[1]).strip()
            
            if company_name not in company_map:
                print(f"  ⚠️  Empresa desconhecida: {company_name}")
                continue
            
            company_id = company_map[company_name]
            
            data = {
                "total_posts": int(row[2]) if row[2] else 0,
                "total_interacoes": int(row[3]) if row[3] else 0,
                "engagement_medio": float(row[4]) if row[4] else 0,
                "alcance_total": int(row[5]) if row[5] else 0,
                "impressoes_total": int(row[6]) if row[6] else 0,
                "seguidores": {
                    "instagram": int(row[7]) if row[7] else 0,
                    "facebook": int(row[8]) if row[8] else 0,
                    "youtube": int(row[9]) if row[9] else 0,
                    "twitter": int(row[10]) if row[10] else 0,
                    "linkedin": int(row[11]) if row[11] else 0,
                    "tiktok": int(row[12]) if row[12] else 0,
                    "threads": int(row[13]) if row[13] else 0,
                }
            }
            
            if insert_snapshot(cursor, company_id, snapshot_date, 'metricool_social', 'metricool', data):
                count += 1
                print(f"  ✓ {snapshot_date.strftime('%Y-%m-%d')} - {company_name}: {data['total_posts']} posts")
        
        except Exception as e:
            print(f"  ✗ Erro na linha: {row[0]} - {e}")
    
    return count

def import_cademi(wb, cursor):
    """Importa dados da aba Cademi Cursos"""
    print("\n📊 Importando Cademi Cursos...")
    
    if "Cademi Cursos" not in wb.sheetnames:
        print("⚠️  Aba 'Cademi Cursos' não encontrada")
        return 0
    
    ws = wb["Cademi Cursos"]
    count = 0
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0] or row[0] == "":
            continue
        
        try:
            snapshot_date = datetime.strptime(str(row[0]), "%Y-%m-%d")
            
            data = {
                "total_alunos": int(row[1]) if row[1] else 0,
                "alunos_ativos": int(row[2]) if row[2] else 0,
                "alunos_inativos": int(row[3]) if row[3] else 0,
                "total_cursos": int(row[4]) if row[4] else 0,
                "taxa_ativacao": float(row[5]) if row[5] else 0,
            }
            
            if insert_snapshot(cursor, 4, snapshot_date, 'cademi_courses', 'cademi', data):
                count += 1
                print(f"  ✓ {snapshot_date.strftime('%Y-%m-%d')}: {data['total_alunos']} alunos")
        
        except Exception as e:
            print(f"  ✗ Erro na linha: {row[0]} - {e}")
    
    return count

def main():
    if len(sys.argv) < 2:
        print("❌ Uso: python3 import_historical_data.py <caminho_para_planilha.xlsx>")
        sys.exit(1)
    
    excel_file = sys.argv[1]
    
    if not os.path.exists(excel_file):
        print(f"❌ Arquivo não encontrado: {excel_file}")
        sys.exit(1)
    
    print(f"📂 Carregando planilha: {excel_file}")
    
    try:
        wb = load_workbook(excel_file, data_only=True)
        print(f"✅ Planilha carregada com sucesso")
        print(f"   Abas encontradas: {', '.join(wb.sheetnames)}")
    except Exception as e:
        print(f"❌ Erro ao carregar planilha: {e}")
        sys.exit(1)
    
    # Conectar ao banco de dados
    try:
        connection = get_db_connection()
        cursor = connection.cursor()
    except Exception as e:
        print(f"❌ Erro ao conectar ao banco: {e}")
        sys.exit(1)
    
    # Importar dados de cada aba
    total_imported = 0
    
    total_imported += import_blue_consult(wb, cursor)
    total_imported += import_tokeniza_academy(wb, cursor)
    total_imported += import_metricool(wb, cursor)
    total_imported += import_cademi(wb, cursor)
    
    # Commit das alterações
    try:
        connection.commit()
        print(f"\n✅ Importação concluída com sucesso!")
        print(f"   Total de snapshots importados: {total_imported}")
    except Error as e:
        print(f"\n❌ Erro ao salvar alterações: {e}")
        connection.rollback()
    finally:
        cursor.close()
        connection.close()
        print("🔒 Conexão com banco de dados fechada")

if __name__ == "__main__":
    main()
