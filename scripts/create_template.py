#!/usr/bin/env python3
"""
Script para criar planilha modelo Excel para importação de dados históricos de KPIs
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

def create_template():
    wb = Workbook()
    
    # Remover sheet padrão
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Criar sheets para cada tipo de snapshot
    create_blue_consult_sheet(wb)
    create_tokeniza_academy_sheet(wb)
    create_metricool_sheet(wb)
    create_cademi_sheet(wb)
    create_instructions_sheet(wb)
    
    # Salvar arquivo
    output_file = '/home/ubuntu/kpi-dashboard/KPI_Import_Template.xlsx'
    wb.save(output_file)
    print(f"✅ Planilha modelo criada: {output_file}")
    return output_file

def style_header(ws, row=1):
    """Aplica estilo ao cabeçalho"""
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    for cell in ws[row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

def auto_adjust_columns(ws):
    """Ajusta largura das colunas automaticamente"""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

def create_blue_consult_sheet(wb):
    """Sheet para Blue Consult (Pipedrive + Nibo)"""
    ws = wb.create_sheet("Blue Consult")
    
    # Cabeçalhos
    headers = [
        "data",
        "faturamento_mensal",
        "novos_clientes",
        "clientes_implantacao",
        "taxa_conversao",
        "receitas_nibo",
        "despesas_nibo",
        "saldo_nibo"
    ]
    
    ws.append(headers)
    style_header(ws)
    
    # Adicionar linha de exemplo
    example_date = datetime.now() - timedelta(days=30)
    ws.append([
        example_date.strftime("%Y-%m-%d"),
        "180000.00",
        "12",
        "61",
        "89.8",
        "17800.00",
        "246300.00",
        "-228600.00"
    ])
    
    # Adicionar mais 6 linhas vazias para preenchimento
    for i in range(6):
        ws.append([""] * len(headers))
    
    auto_adjust_columns(ws)
    
    # Adicionar nota
    ws['A10'] = "INSTRUÇÕES:"
    ws['A10'].font = Font(bold=True, size=11)
    ws['A11'] = "• data: Formato YYYY-MM-DD (ex: 2024-10-01)"
    ws['A12'] = "• faturamento_mensal: Valor em reais sem símbolo (ex: 180000.00)"
    ws['A13'] = "• novos_clientes: Número inteiro de clientes novos"
    ws['A14'] = "• clientes_implantacao: Número de clientes em implantação"
    ws['A15'] = "• taxa_conversao: Percentual sem símbolo % (ex: 89.8)"
    ws['A16'] = "• receitas_nibo: Receitas do Nibo em reais"
    ws['A17'] = "• despesas_nibo: Despesas do Nibo em reais"
    ws['A18'] = "• saldo_nibo: Saldo (receitas - despesas)"

def create_tokeniza_academy_sheet(wb):
    """Sheet para Tokeniza Academy (Discord + Cademi)"""
    ws = wb.create_sheet("Tokeniza Academy")
    
    headers = [
        "data",
        "total_membros_discord",
        "membros_online",
        "novos_membros_7d",
        "novos_membros_30d",
        "total_alunos_cademi",
        "alunos_ativos",
        "total_cursos"
    ]
    
    ws.append(headers)
    style_header(ws)
    
    # Exemplo
    example_date = datetime.now() - timedelta(days=30)
    ws.append([
        example_date.strftime("%Y-%m-%d"),
        "1854",
        "154",
        "5",
        "6",
        "450",
        "320",
        "8"
    ])
    
    for i in range(6):
        ws.append([""] * len(headers))
    
    auto_adjust_columns(ws)
    
    ws['A10'] = "INSTRUÇÕES:"
    ws['A10'].font = Font(bold=True, size=11)
    ws['A11'] = "• data: Formato YYYY-MM-DD"
    ws['A12'] = "• total_membros_discord: Total de membros no Discord"
    ws['A13'] = "• membros_online: Membros online no momento"
    ws['A14'] = "• novos_membros_7d: Novos membros nos últimos 7 dias"
    ws['A15'] = "• novos_membros_30d: Novos membros nos últimos 30 dias"
    ws['A16'] = "• total_alunos_cademi: Total de alunos cadastrados"
    ws['A17'] = "• alunos_ativos: Alunos com acesso ativo"
    ws['A18'] = "• total_cursos: Número de cursos disponíveis"

def create_metricool_sheet(wb):
    """Sheet para Metricool (Redes Sociais)"""
    ws = wb.create_sheet("Redes Sociais")
    
    headers = [
        "data",
        "empresa",
        "total_posts",
        "total_interacoes",
        "engagement_medio",
        "alcance_total",
        "impressoes_total",
        "seguidores_instagram",
        "seguidores_facebook",
        "seguidores_youtube",
        "seguidores_twitter",
        "seguidores_linkedin",
        "seguidores_tiktok",
        "seguidores_threads"
    ]
    
    ws.append(headers)
    style_header(ws)
    
    # Exemplos para cada empresa
    example_date = datetime.now() - timedelta(days=30)
    companies = [
        ("Blue Consult", "61", "363", "2.09", "6600", "95400", "14200", "1", "202", "0", "0", "0", "0"),
        ("Tokeniza", "61", "363", "2.09", "6600", "95400", "14200", "1", "202", "0", "0", "0", "0"),
        ("Tokeniza Academy", "180", "229", "0.1", "4940", "8500", "1200", "50", "100", "0", "0", "0", "0"),
        ("Mychel Mendes", "725", "9400", "0.57", "99200", "190300", "52800", "1", "97200", "0", "0", "300", "0"),
    ]
    
    for company_data in companies:
        row = [example_date.strftime("%Y-%m-%d")] + list(company_data)
        ws.append(row)
    
    for i in range(3):
        ws.append([""] * len(headers))
    
    auto_adjust_columns(ws)
    
    ws['A12'] = "INSTRUÇÕES:"
    ws['A12'].font = Font(bold=True, size=11)
    ws['A13'] = "• data: Formato YYYY-MM-DD"
    ws['A14'] = "• empresa: Nome da empresa (Blue Consult, Tokeniza, Tokeniza Academy, Mychel Mendes)"
    ws['A15'] = "• total_posts: Número de posts publicados"
    ws['A16'] = "• total_interacoes: Soma de curtidas, comentários, compartilhamentos"
    ws['A17'] = "• engagement_medio: Taxa de engajamento em % sem símbolo (ex: 2.09)"
    ws['A18'] = "• alcance_total: Número de pessoas alcançadas"
    ws['A19'] = "• impressoes_total: Número total de impressões"
    ws['A20'] = "• seguidores_*: Número de seguidores em cada rede social"

def create_cademi_sheet(wb):
    """Sheet para Cademi (Cursos)"""
    ws = wb.create_sheet("Cademi Cursos")
    
    headers = [
        "data",
        "total_alunos",
        "alunos_ativos",
        "alunos_inativos",
        "total_cursos",
        "taxa_ativacao"
    ]
    
    ws.append(headers)
    style_header(ws)
    
    # Exemplo
    example_date = datetime.now() - timedelta(days=30)
    ws.append([
        example_date.strftime("%Y-%m-%d"),
        "450",
        "320",
        "130",
        "8",
        "71.1"
    ])
    
    for i in range(6):
        ws.append([""] * len(headers))
    
    auto_adjust_columns(ws)
    
    ws['A10'] = "INSTRUÇÕES:"
    ws['A10'].font = Font(bold=True, size=11)
    ws['A11'] = "• data: Formato YYYY-MM-DD"
    ws['A12'] = "• total_alunos: Total de alunos cadastrados"
    ws['A13'] = "• alunos_ativos: Alunos com acesso ativo aos cursos"
    ws['A14'] = "• alunos_inativos: Alunos sem acesso ativo"
    ws['A15'] = "• total_cursos: Número de cursos disponíveis"
    ws['A16'] = "• taxa_ativacao: Percentual de alunos ativos (ex: 71.1)"

def create_instructions_sheet(wb):
    """Sheet com instruções gerais"""
    ws = wb.create_sheet("📋 INSTRUÇÕES", 0)  # Primeira aba
    
    ws.column_dimensions['A'].width = 100
    
    instructions = [
        ("PLANILHA MODELO PARA IMPORTAÇÃO DE DADOS HISTÓRICOS DE KPIs", "title"),
        ("", ""),
        ("📌 COMO USAR ESTA PLANILHA:", "header"),
        ("", ""),
        ("1. Preencha cada aba com os dados históricos correspondentes", "text"),
        ("2. Respeite o formato de data: YYYY-MM-DD (ex: 2024-10-01)", "text"),
        ("3. Use números sem símbolos de moeda ou porcentagem", "text"),
        ("4. Valores decimais devem usar ponto (.) e não vírgula (,)", "text"),
        ("5. Não altere os nomes das colunas (primeira linha)", "text"),
        ("6. Você pode adicionar quantas linhas quiser em cada aba", "text"),
        ("7. Salve o arquivo após preencher", "text"),
        ("8. Envie o arquivo preenchido para importação", "text"),
        ("", ""),
        ("📊 ABAS DISPONÍVEIS:", "header"),
        ("", ""),
        ("• Blue Consult: Dados de vendas (Pipedrive) e financeiro (Nibo)", "text"),
        ("• Tokeniza Academy: Dados do Discord e plataforma Cademi", "text"),
        ("• Redes Sociais: Métricas de todas as redes sociais (Metricool)", "text"),
        ("• Cademi Cursos: Dados detalhados da plataforma de cursos", "text"),
        ("", ""),
        ("⚠️ IMPORTANTE:", "header"),
        ("", ""),
        ("• Cada linha representa um snapshot diário (uma data específica)", "text"),
        ("• Recomendado preencher dados de pelo menos 30 dias para comparações mensais", "text"),
        ("• Dados mais antigos permitem análises de tendências mais precisas", "text"),
        ("• Se não tiver um dado específico, deixe a célula vazia", "text"),
        ("", ""),
        ("💡 DICAS:", "header"),
        ("", ""),
        ("• Comece preenchendo os dados mais recentes e vá voltando no tempo", "text"),
        ("• Use os exemplos fornecidos em cada aba como referência", "text"),
        ("• Mantenha consistência nos formatos de data e números", "text"),
        ("• Após a primeira importação, o sistema coletará dados automaticamente", "text"),
        ("", ""),
        ("", ""),
        ("Criado em: " + datetime.now().strftime("%d/%m/%Y %H:%M"), "footer"),
    ]
    
    row = 1
    for text, style_type in instructions:
        cell = ws.cell(row=row, column=1, value=text)
        
        if style_type == "title":
            cell.font = Font(bold=True, size=16, color="1F4E78")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        elif style_type == "header":
            cell.font = Font(bold=True, size=13, color="2E75B6")
        elif style_type == "text":
            cell.font = Font(size=11)
            cell.alignment = Alignment(wrap_text=True)
        elif style_type == "footer":
            cell.font = Font(size=9, italic=True, color="7F7F7F")
            cell.alignment = Alignment(horizontal="right")
        
        row += 1
    
    # Ajustar altura das linhas
    for row_num in range(1, row):
        ws.row_dimensions[row_num].height = 20

if __name__ == "__main__":
    create_template()
